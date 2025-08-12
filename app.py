from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
import sqlite3
from datetime import datetime, date
import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment, numbers
import uuid

app = Flask(__name__)
DATABASE = 'pedidos.db'

def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()

    conn.execute('''
        CREATE TABLE IF NOT EXISTS productos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            precio REAL NOT NULL
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS pedidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            total REAL NOT NULL,
            estado TEXT NOT NULL
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS detalle_pedidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pedido_id INTEGER NOT NULL,
            producto_id INTEGER NOT NULL,
            cantidad INTEGER NOT NULL,
            subtotal REAL NOT NULL,
            FOREIGN KEY (pedido_id) REFERENCES pedidos(id),
            FOREIGN KEY (producto_id) REFERENCES productos(id)
        )
    ''')

    conn.commit()
    conn.close()
    
init_db()

def agregar_columna_archivado():
    conn = get_db_connection()
    try:
        conn.execute('ALTER TABLE pedidos ADD COLUMN archivado INTEGER DEFAULT 0')
        conn.commit()
    except sqlite3.OperationalError:
        pass
    conn.close()

def agregar_columna_turno():
    conn = get_db_connection()
    try:
        conn.execute('ALTER TABLE pedidos ADD COLUMN turno_id TEXT DEFAULT NULL')
        conn.commit()
    except sqlite3.OperationalError:
        pass
    conn.close()

agregar_columna_archivado()
agregar_columna_turno()

app.secret_key = "supersecret"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/agregar', methods=['GET', 'POST'])
def agregar_pedido():
    conn = get_db_connection()
    productos = conn.execute('SELECT * FROM productos').fetchall()
    
    if not productos:
        flash("Primero debes agregar productos antes de poder crear un pedido.", "warning")
        return redirect(url_for('editar_productos'))

    if request.method == 'POST':
        cantidades = {str(p['id']): int(request.form.get(f"cantidad_{p['id']}", 0)) for p in productos}
        
        seleccionados = [(p['id'], p['precio'], cantidades[str(p['id'])]) for p in productos if cantidades[str(p['id'])] > 0]

        if not seleccionados:
            flash("Debe seleccionar al menos un producto", "warning")
            return redirect(url_for('agregar_pedido'))

        total = sum(precio * cant for _, precio, cant in seleccionados)

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur = conn.cursor()
        cur.execute('INSERT INTO pedidos (fecha, total, estado) VALUES (?, ?, ?)', (fecha, total, "Pendiente"))
        pedido_id = cur.lastrowid

        for prod_id, precio, cant in seleccionados:
            subtotal = precio * cant
            cur.execute('INSERT INTO detalle_pedidos (pedido_id, producto_id, cantidad, subtotal) VALUES (?, ?, ?, ?)',
                        (pedido_id, prod_id, cant, subtotal))

        conn.commit()
        conn.close()

        flash(f"El pedido se realizó correctamente, el total es de ${total:.2f}", "success")
        return redirect(url_for('pedidos_pendientes'))

    conn.close()
    return render_template('agregar_pedido.html', productos=productos)

@app.route('/pendientes')
def pedidos_pendientes():
    conn = get_db_connection()

    pedidos = conn.execute('SELECT * FROM pedidos WHERE estado = "Pendiente"').fetchall()

    lista_pedidos = []
    for pedido in pedidos:
        detalles = conn.execute('''
            SELECT productos.nombre, detalle_pedidos.cantidad, detalle_pedidos.subtotal
            FROM detalle_pedidos
            LEFT JOIN productos ON productos.id = detalle_pedidos.producto_id
            WHERE detalle_pedidos.pedido_id = ?
        ''', (pedido['id'],)).fetchall()

        productos_eliminados = any(d['nombre'] is None for d in detalles)

        lista_pedidos.append({
            'id': pedido['id'],
            'fecha': pedido['fecha'],
            'total': pedido['total'],
            'estado': pedido['estado'],
            'detalles': detalles,
            'detalles_vacios': len(detalles) == 0,
            'productos_eliminados': productos_eliminados
        })

    conn.close()
    return render_template('pendientes.html', pedidos=lista_pedidos)

@app.route('/pedidos/completar/<int:id>')
def completar_pedido(id):
    conn = get_db_connection()
    pedido = conn.execute('SELECT * FROM pedidos WHERE id = ? AND estado = "Pendiente"', (id,)).fetchone()
    if pedido is None:
        conn.close()
        flash("Pedido no encontrado o ya completado.", "warning")
        return redirect(url_for('pedidos_pendientes'))

    conn.execute('UPDATE pedidos SET estado = "Completado" WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    flash(f"Pedido #{id} marcado como completado.", "success")
    return redirect(url_for('pedidos_pendientes'))

@app.route('/pedidos/editar/<int:id>', methods=['GET', 'POST'])
def editar_pedido(id):
    conn = get_db_connection()

    productos = conn.execute('SELECT * FROM productos').fetchall()

    detalles_actuales = conn.execute('''
        SELECT productos.nombre, detalle_pedidos.producto_id, detalle_pedidos.cantidad
        FROM detalle_pedidos
        LEFT JOIN productos ON productos.id = detalle_pedidos.producto_id
        WHERE detalle_pedidos.pedido_id = ?
    ''', (id,)).fetchall()

    cantidades_actuales = {}
    productos_eliminados = False
    for d in detalles_actuales:
        if d['nombre'] is None:
            productos_eliminados = True
        cantidades_actuales[d['producto_id']] = d['cantidad']

    if request.method == 'POST':
        cantidades = {str(p['id']): int(request.form.get(f"cantidad_{p['id']}", 0)) for p in productos}
        seleccionados = [(p['id'], p['precio'], cantidades[str(p['id'])]) for p in productos if cantidades[str(p['id'])] > 0]

        if not seleccionados:
            flash("Debe seleccionar al menos un producto", "warning")
            return redirect(url_for('editar_pedido', id=id))

        total = sum(precio * cant for _, precio, cant in seleccionados)

        conn.execute('DELETE FROM detalle_pedidos WHERE pedido_id = ?', (id,))

        for prod_id, precio, cant in seleccionados:
            subtotal = precio * cant
            conn.execute('INSERT INTO detalle_pedidos (pedido_id, producto_id, cantidad, subtotal) VALUES (?, ?, ?, ?)',
                         (id, prod_id, cant, subtotal))

        conn.execute('UPDATE pedidos SET total = ? WHERE id = ?', (total, id))

        conn.commit()
        conn.close()
        flash(f"Pedido #{id} editado correctamente. Nuevo total: ${total:.2f}", "success")
        return redirect(url_for('pedidos_pendientes'))

    conn.close()
    return render_template('editar_pedido.html', id=id, productos=productos, cantidades=cantidades_actuales, productos_eliminados=productos_eliminados)


@app.route('/completados')
def pedidos_completados():
    conn = get_db_connection()
    pedidos = conn.execute('SELECT * FROM pedidos WHERE estado = "Completado" AND archivado = 0').fetchall()

    lista_pedidos = []
    for pedido in pedidos:
        detalles = conn.execute('''
            SELECT productos.nombre, detalle_pedidos.cantidad, detalle_pedidos.subtotal
            FROM detalle_pedidos
            LEFT JOIN productos ON productos.id = detalle_pedidos.producto_id
            WHERE detalle_pedidos.pedido_id = ?
        ''', (pedido['id'],)).fetchall()

        productos_eliminados = any(d['nombre'] is None for d in detalles)

        lista_pedidos.append({
            'id': pedido['id'],
            'fecha': pedido['fecha'],
            'total': pedido['total'],
            'estado': pedido['estado'],
            'detalles': detalles,
            'detalles_vacios': len(detalles) == 0,
            'productos_eliminados': productos_eliminados
        })

    conn.close()
    return render_template('completados.html', pedidos=lista_pedidos)

@app.route('/productos', methods=['GET', 'POST'])
def editar_productos():
    conn = get_db_connection()
    
    if request.method == 'POST':
        nombre = request.form['nombre']
        precio = request.form['precio']
        
        conn.execute('INSERT INTO productos (nombre, precio) VALUES (?, ?)',
                     (nombre, precio))
        conn.commit()
        conn.close()
        return redirect(url_for('editar_productos'))
    
    productos = conn.execute('SELECT * FROM productos').fetchall()
    conn.close()
    return render_template('productos.html', productos=productos)

@app.route('/productos/editar/<int:id>', methods=['GET', 'POST'])
def editar_producto(id):
    conn = get_db_connection()
    producto = conn.execute('SELECT * FROM productos WHERE id = ?', (id,)).fetchone()

    if not producto:
        conn.close()
        flash("Producto no encontrado", "warning")
        return redirect(url_for('editar_productos'))

    if request.method == 'POST':
        nombre = request.form['nombre']
        precio = request.form['precio']
        conn.execute('UPDATE productos SET nombre = ?, precio = ? WHERE id = ?', (nombre, precio, id))
        conn.commit()
        conn.close()
        flash("Producto actualizado correctamente", "success")
        return redirect(url_for('editar_productos'))

    conn.close()
    return render_template('editar_productos.html', producto=producto)

@app.route('/productos/eliminar/<int:id>')
def eliminar_producto(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM productos WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    flash("Producto eliminado correctamente", "success")
    return redirect(url_for('editar_productos'))

@app.route('/cerrar')
def cerrar_dia():
    hoy = date.today().strftime('%Y-%m-%d')

    conn = get_db_connection()
    pedidos_completados = conn.execute(
        'SELECT * FROM pedidos WHERE estado = "Completado" AND archivado = 0 AND fecha LIKE ?', (hoy + '%',)
    ).fetchall()

    total_dia = sum(p['total'] for p in pedidos_completados)

    conn.close()

    return render_template('cerrar.html', pedidos=pedidos_completados, total=total_dia, fecha=hoy)

@app.route('/cerrar/confirmar', methods=['POST'])
def confirmar_cierre_dia():
    hoy = date.today().strftime('%Y-%m-%d')
    cierre_id = str(uuid.uuid4())  # ID único para este turno/cierre

    conn = get_db_connection()

    # Pedidos completados sin turno_id asignado del día actual
    pedidos_completados = conn.execute(
        'SELECT * FROM pedidos WHERE estado = "Completado" AND (turno_id IS NULL OR turno_id = "") AND fecha LIKE ?', (hoy + '%',)
    ).fetchall()

    # Asignar turno_id a esos pedidos
    conn.execute(
        'UPDATE pedidos SET turno_id = ? WHERE estado = "Completado" AND (turno_id IS NULL OR turno_id = "") AND fecha LIKE ?',
        (cierre_id, hoy + '%')
    )
    conn.commit()

    detalles = []
    for p in pedidos_completados:
        det = conn.execute(
            '''SELECT productos.nombre, detalle_pedidos.cantidad, detalle_pedidos.subtotal
               FROM detalle_pedidos
               JOIN productos ON productos.id = detalle_pedidos.producto_id
               WHERE detalle_pedidos.pedido_id = ?''', (p['id'],)
        ).fetchall()
        detalles.append((p, det))

    session['cierre_dia'] = {
        'fecha': hoy,
        'turno_id': cierre_id,
        'pedidos': [
            {
                'id': p['id'],
                'fecha': p['fecha'],
                'total': p['total'],
                'detalles': [{'nombre': d['nombre'], 'cantidad': d['cantidad'], 'subtotal': d['subtotal']} for d in det]
            } for p, det in detalles
        ]
    }

    # Aquí archivamos solo pedidos de este turno
    conn.execute('UPDATE pedidos SET archivado = 1 WHERE turno_id = ?', (cierre_id,))
    
    # Limpiamos pedidos pendientes (opcional según lógica)
    conn.execute('DELETE FROM detalle_pedidos WHERE pedido_id IN (SELECT id FROM pedidos WHERE estado = "Pendiente")')
    conn.execute('DELETE FROM pedidos WHERE estado = "Pendiente"')
    conn.commit()
    conn.close()

    return render_template('cierre_exitoso.html', fecha=hoy)

@app.route('/cerrar/descargar_excel')
def descargar_excel():
    if 'cierre_dia' not in session:
        flash("No hay datos para descargar. Debe cerrar el día primero.", "warning")
        return redirect(url_for('index'))

    cierre = session['cierre_dia']
    hoy = cierre['fecha']
    turno_id = cierre['turno_id']
    pedidos = cierre['pedidos']

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        rows = []
        for pedido in pedidos:
            for d in pedido['detalles']:
                rows.append({
                    'Pedido Nro': pedido['id'],
                    'Fecha': pd.to_datetime(pedido['fecha']),
                    'Producto': d['nombre'],
                    'Cantidad': d['cantidad'],
                    'Subtotal': d['subtotal'],
                    'Total Pedido': pedido['total']
                })

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Pedidos Completados', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Pedidos Completados']
        total_general = sum(p['total'] for p in pedidos)

        last_row = len(df) + 2
        worksheet.cell(row=last_row, column=5, value="TOTAL GENERAL:")
        worksheet.cell(row=last_row, column=6, value=total_general)
        
        column_widths = {
    'A': 12,
    'B': 12,
    'C': 25,
    'D': 12,
    'E': 12,
    'F': 15
    }
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        # Definir estilos comunes
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Aplicar estilos a todo el rango con datos
        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = align_center

                # Formatear fechas
                if cell.column_letter == 'B':  # Columna Fecha
                    cell.number_format = 'DD/MM/YYYY'

                # Formatear precios con signo $
                if cell.column_letter in ['E', 'F']:  # Subtotal y Total Pedido
                    cell.number_format = '"$"#,##0.00'

        # Formato y alineación para fila de total general
        cell_total_label = worksheet.cell(row=last_row, column=5)
        cell_total_value = worksheet.cell(row=last_row, column=6)
        cell_total_label.alignment = align_center
        cell_total_value.alignment = align_center
        cell_total_value.number_format = '"$"#,##0.00'

    output.seek(0)

    session.pop('cierre_dia')

    return send_file(
        output,
        as_attachment=True,
        download_name=f"Cierre_Turno_{turno_id[:8]}_{hoy}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == "__main__":
    app.run(debug=True)